#Import Selenium Library
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait

#Import Others Library
import os
import wget
import time
from selenium.webdriver.support.ui import Select
from datetime import datetime

from selenium.webdriver.support.select import Select
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
import pandas as pd
import xlrd
import datetime
import openpyxl
import pyautogui
from datetime import timedelta
from selenium.common.exceptions import TimeoutException

#Login to SIPMEN Account
driver = webdriver.Chrome(ChromeDriverManager().install())
driver.get("https://app.32net.id/link/go/sipmen32/") 

email = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "input[id='email']")))

email.clear()
email.send_keys("*********@***.**.id")

password = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "input[id='password']")))

password.clear()
password.send_keys("**********")

masuk = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "button[type='submit']"))).click()

time.sleep(1)

distribusi = driver.find_element(By.CSS_SELECTOR, ".z-20:nth-child(1) .relative:nth-child(4) .ml-4").click()

ExcelPath = 'C:/Users/MYPC PRO L7/Downloads/monitor.xlsx'
df = pd.read_excel(ExcelPath)
wb = openpyxl.load_workbook(ExcelPath)
ws = wb.worksheets[0]
maxrow = ws.max_row

driver.get('https://app.32net.id/link/go/sipmen32/pengambilan')
time.sleep(0.5)

i = 2

while i <= len(ws['A']):
    id_batch = ws['B'+str(i)].value
    slss = ws['C'+str(i)].value
    op_entry = ws['A'+str(i)].value
    a_tgl = ws['D'+str(i)].value
    
    driver.get('https://app.32net.id/link/go/sipmen32/pengambilan')
    
    try:
    
        time.sleep(3)
        #id_batch
        driver.find_element(By.XPATH, '//*[@id="id_batch"]').click()
        pyautogui.typewrite(str(id_batch))
        time.sleep(0.5)
        # slss
        driver.find_element(
        By.XPATH, '/html/body/div/div[2]/main/div/form/div/label[2]/span[2]/span[1]/span').click()
        pyautogui.typewrite(str(slss))
        time.sleep(0.5)
        pyautogui.press('enter')
        time.sleep(0.5)
        # op_entry
        driver.find_element(By.XPATH, '//*[@id="select2-op_entry-container"]').click()
        pyautogui.typewrite(op_entry)
        time.sleep(0.5)
        pyautogui.press('enter') 
        time.sleep(1)
        # tgl
        tgl = driver.find_element(By.XPATH, '//*[@id="a_tgl"]').click()
        pyautogui.typewrite(a_tgl)
        time.sleep(0.5)
        #submit
        submit = driver.find_element("xpath", '/html/body/div/div[2]/main/div/form/div/div/button').click()
        
    except TimeoutException:
        print("form hilang")
        pass
    
    time.sleep(1)
    i=i+1

print("AUTOMASI SIPMENKU BERHASIL :) TERIMAKASIH TUHAN")
